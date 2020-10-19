from datetime import datetime, timedelta
from textblob import TextBlob
import mysql.connector
from collections import OrderedDict
from .fusioncharts import FusionCharts
from datetime import datetime
import json
from requests.auth import HTTPDigestAuth
import tweepy
import pandas as pd
import datetime
from os import environ
from django.db import models

import time           # time date libarary
import json           # Json paring

from django.shortcuts import render

import io        # seting eniromental variable
import os        # setting enivromental varible
import xlwt
from django.http import HttpResponse
import sys

# my scraping code start here

import requests
import urllib.request
import time
from bs4 import BeautifulSoup
import re
import sqlite3
import sys
# from firebase import firebase
# firebase = firebase.FirebaseApplication(
#     'https://twitionary.firebaseio.com/', None)
from datetime import timedelta
from datetime import datetime
import openpyxl as xl

# newshunt

# json object to go to news page

# used for chart
consumer_key = "iJFZnuM0YHqwvFilNUBSVkzJU"
consumer_secret = "412n9RVFyUc4lRH3RWBU4kRT1lz5NHWg81d6FEoMPQEvYJPRio"
access_token = "813456180-jXG4M0Kpc80UjJF4bhwA0z9Bx8aZfAht4veyxSgc"
access_token_secret = "Cvh3gsUS5Y4HpyfD1UdlMLfpxlqa47iYFo3vxogpP6blR"
auth = tweepy.OAuthHandler(consumer_key, consumer_secret)
auth.set_access_token(access_token, access_token_secret)
api = tweepy.API(auth,wait_on_rate_limit=True)

# SQL Connection String strats
mydb = mysql.connector.connect(
    host="104.155.148.67",
    # # host="localhost",
    database="twitter",
    user="twitteruser",
    passwd="irfansyed",
    # host="localhost",
    # database="twitter",
    # user="root",
    # passwd="",
)
cursor = mydb.cursor()
# countcursor = mydb.cursor(buffered=True)
# SQL Connection String Ends
# getNews When Get News Button Clicked


def getNewsButton(request):

    posts = []
    date1 = request.GET.get("date1")
    date2 = request.GET.get("date2")
    country = request.GET.get("contry")
    # lang=request.GET.get("lang")
  #  if lang=='Urdu':
    # posts=getNewsUrdu(date,country)
    # else:
    posts = getNewsEnglish(date1, date2, country)
    countries = getCountryNames()

    context = {

        'posts': posts,

        'country': countries,
        'date1': date1,
        'date2': date2

    }

    return render(request, 'blog/news.html', context)


def getNewsEnglish(newsdate1, newsdate2, country):

    if newsdate1 == newsdate2:
        datee = datetime.today().strftime('%Y-%m-%d')
        datee2 = datetime.today() + timedelta(days=1)
        datee2 = datee2.strftime('%Y-%m-%d')
        newsdate1 = datee
        newsdate2 = datee2

    posts = []
    df = pd.read_excel('dictionary.xlsx')
    mylist = df['words'].tolist()

    r = re.compile('|'.join([r'\b%s\b' % w for w in mylist]), flags=re.I)

    url = "https://newshunt.io/getBetweenDates/"+newsdate1+'/'+newsdate2
    if country != 'World':
        url = "https://newshunt.io/getBetweenDates/"+newsdate1+'/'+newsdate2+'/'+country

    print(url)

    myResponse = requests.get(url, verify=True)
    if(myResponse.ok):
        jData = json.loads(myResponse.content)

        jData = jData['news']
        for key in jData:
            listt = r.findall(key['description'])

            if listt:
                header = re.sub('^A-Za-z0-9]+ +', ' ', key['title'])
                prgh = re.sub('^A-Za-z0-9]+ +', ' ', key['description'])
                header = re.sub("\s\s+", " ", header)
                prgh = re.sub("\s\s+", " ", prgh)
                sentiments = sentiment(prgh)
                posts.append({

                    'href': key['url'],
                    'img':  key['media'],
                    'header': header,
                    'prgh': prgh,
                    'date': key['pub_date'],
                    'News': key['source'],
                    'country': key['country'],
                    'sentiment': sentiments,
                    'keyWord': listt,

                })

    return posts

def getNewsUrdu(newsdate, country):

    posts = []
    df = pd.read_excel('dictionary.xlsx')
    mylist = df['words'].tolist()
    r = re.compile('|'.join([r'\b%s\b' % w for w in mylist]), flags=re.I)

    url = "https://newshunt.io/getDateNews/"+newsdate

    count = []
    myResponse = requests.get(url, verify=True)
    if(myResponse.ok):
        jData = json.loads(myResponse.content)

        jData = jData['news']
        for key in jData:
            if key['country'] not in count:
                count.append(key['country'])
            listt = r.findall(key['description'])
            if country == key['country'] and key['lang'] == 'urdu':
                # if listt:
                header = re.sub('^A-Za-z0-9]+ +', ' ', key['title'])
                prgh = re.sub('^A-Za-z0-9]+ +', ' ', key['description'])
                header = re.sub("\s\s+", " ", header)
                prgh = re.sub("\s\s+", " ", prgh)

                posts.append({

                    'href': key['url'],
                    'img':  key['media'],
                    'header': header,
                    'prgh': prgh,
                    'date': newsdate,
                    'News': key['source'],
                    'words': key['country'],
                    'sentiment': "NA",

                })
            elif country == "World" and key['lang'] == 'urdu':
                # if listt:
                header = re.sub('^A-Za-z0-9]+ +', ' ', key['title'])
                prgh = re.sub('^A-Za-z0-9]+ +', ' ', key['description'])
                header = re.sub("\s\s+", " ", header)
                prgh = re.sub("\s\s+", " ", prgh)

                posts.append({

                    'href': key['url'],
                    'img':  key['media'],
                    'header': header,
                    'prgh': prgh,
                    'date': newsdate,
                    'News': key['source'],
                    'words': key['country'],
                    'sentiment': "NA",

                })

    return posts


countries = []


def news(request):

    datee = datetime.today().strftime('%Y-%m-%d')

    posts = []
    posts = getNewsEnglish(datee, datee, 'pakistan')

    countries = getCountryNames()

    context = {

        'posts': posts,
        'date1': datee,
        'date2': datee,
        'country': countries

    }

    # (request,the blog i am requestin,my json object)
    return render(request, 'blog/news.html', context)
    # return render(request, 'blog/news.html', {'tital': 'news'})


def getCountryNames():
    url = "https://newshunt.io/getCountriesIrfan"

    country = []
    myResponse = requests.get(url, verify=True)
    if(myResponse.ok):
        jData = json.loads(myResponse.content)
        jData = jData['country']

        for key in jData:
            country.append(key["name"])

    return country


def login(request):

    context = {



    }

    # (request,the blog i am requestin,my json object)
    return render(request, 'blog/login.html', context)


def home(request):
    return render(request, 'blog/home.html', {'tital': 'Home'})

def index (request):
    # def Followers(username):

    return render(request, 'blog/index.html')

# Twiteer scraping
def chart(request):

    posts = []
    mydb._open_connection()
    sqlite_select_query = """SELECT * FROM tbl_twitter ORDER BY id DESC LIMIT 30000"""
    cursor.execute(sqlite_select_query)

    for row in cursor:
        sentiments = sentiment(row[4])
        posts.append({
            'text': row[4],
            'username': row[5],
            'dated': row[2],
            'retweetcount': row[3],
            'location': row[1],
            'urll': row[6],
            'description': row[7],
            'following': row[8],
            'followers': row[9],
            'sentiment': sentiments})
    mydb.commit()
    # show data from 2nd table
    tweets_countq_query = """select COUNT(distinct Username) as username,COUNT(Id),COUNT(distinct location) as location from tbl_twitter order by Id"""
    cursor.execute(tweets_countq_query)
    count_row = cursor.fetchone()
    count_username = count_row[0]
    tweets_count = count_row[1]
    tweets_location = count_row[2]
    wb = xl.load_workbook("dictionary.xlsx", enumerate)
    sheet = wb.worksheets[0]

    tweets_total_keywords = sheet.max_row
    # column_count = sheet.max_column

# start charts from here
    dataSource = OrderedDict()
    # The `chartConfig` dict contains key-value pairs of data for chart attribute
    chartConfig = OrderedDict()
    chartConfig["caption"] = "TWEETS BY SENTIMENT"
    chartConfig["theme"] = "fusion"

    dataSource["chart"] = chartConfig
    dataSource["data"] = []
    dataSource["data"].append({"label": 'Total Tweets', "value": tweets_count})
    dataSource["data"].append(
        {"label": 'Total Users', "value": count_username})
    dataSource["data"].append(
        {"label": 'Total Location', "value": tweets_location})
    dataSource["data"].append(
        {"label": 'Key Words Define', "value": tweets_total_keywords})
    pie3d = FusionCharts("pie3d", "ex2", "100%", "400", "chart-1", "json",
                         # The data is passed as a string in the `dataSource` as parameter.

                         {
                             "chart": {
                                 "caption": "TWEETS BY TYPE",
                                 # "subCaption" : "For a net-worth of $1M",
                                 "showValues": "1",
                                 "showPercentInTooltip": "0",
                                 # "numberPrefix" : "$",
                                 "enableMultiSlicing": "1",
                                 "theme": "fusion"
                             },
                             "data": [{
                                 "label": "Total Tweets",
                                 "value": tweets_count

                             }, {
                                 "label": "Total Users",
                                 "value": count_username
                             }, {
                                 "label": "Total Location",
                                 "value": tweets_location
                             }, {
                                 "label": "Key Words Define",
                                 "value": tweets_total_keywords

                             }]
                         })

    column2D = FusionCharts("column2d", "myFirstChart", "600",
                            "400", "myFirstchart-container", "json", dataSource)

    return render(request, 'blog/translationdetail.html', {
        'posts': posts, 'output': pie3d.render(), 'output1': column2D.render(), 'tweets_count': tweets_count, 'count_username': count_username, 'tweets_location': tweets_location, 'tweets_total_keywords': tweets_total_keywords
    })
#search the twitters 


def sentiment(text):
    analysis = TextBlob(text)
    # set sentiment
    if analysis.sentiment.polarity >= 0:
        return 'positive'
    else:
        return 'negative'

#this is used for sentence analysis
def sentiment_analysis(text):
   
    synt_analysis = TextBlob(text)
    if synt_analysis.sentiment.polarity>=1.0:
        return 'Great'
    elif synt_analysis.sentiment.polarity>=0.7:
        return 'Good'
    elif synt_analysis.sentiment.polarity>=0.0:  
        return 'Neutral'
    elif synt_analysis.sentiment.polarity>=-0.6999998 :  
         return 'Bad'
    elif synt_analysis.sentiment.polarity>=-1.0:  
        return 'Terrible'


# dicinoaty data table

def dictionary_link(request):
    df = pd.read_excel('dictionary.xlsx')
    #items = Laptops.objects.all()
    posts = []
    for index, row in df.iterrows():
        posts.append({
            'words': row['words'],
            'pk': row['pk'],
            })
    context = {
        'items': posts,
        'header': 'Laptops',
    }
    return render(request, 'blog/dictionary.html', context)


def delete_laptop(request):
    pk = request.GET.get("pk")
    df = pd.read_excel('dictionary.xlsx')

    df = df.query("pk != "+pk)
    df.to_excel('dictionary.xlsx')
    df = pd.read_excel('dictionary.xlsx')
    posts = []

    for index, row in df.iterrows():
        posts.append({
            'words': row['words'],
            'pk': row['pk'],

        })

    # items=df
    context = {
        'items': posts,
        'header': 'Laptops',
    }

    return render(request, 'blog/dictionary.html', context)


# from .forms import NameForm
def search_keywords(request):
    
    posts = []
    mydb._open_connection()
    price_lte = request.GET['date1']
    sqlite_select_query = "SELECT * FROM tbl_twitter a JOIN tbl_hashtags b on a.Id=b.twitter_id WHERE b.title= %s ORDER BY a.Id DESC LIMIT 3000"
    cursor.execute(sqlite_select_query,(price_lte,))

    for row in cursor:
        sentiments = sentiment(row[4])
        posts.append({
            'text': row[4],
            'username': row[5],
            'dated': row[2],
            'retweetcount': row[3],
            'location': row[1],
            'urll': row[6],
            'description': row[7],
            'following': row[8],
            'followers': row[9],
            'sentiment': sentiments})
    mydb.commit()
    # show data from 2nd table
    tweets_countq_query = """select COUNT(distinct Username) as username,COUNT(a.Id),COUNT(distinct location) as location from tbl_twitter a JOIN tbl_hashtags b on a.Id=b.twitter_id WHERE b.title= %s ORDER BY a.Id"""
    cursor.execute(tweets_countq_query,(price_lte,))
    count_row = cursor.fetchone()
    count_username = count_row[0]
    tweets_count = count_row[1]
    tweets_location = count_row[2]
    # end query

    wb = xl.load_workbook("dictionary.xlsx", enumerate)
    sheet = wb.worksheets[0]

    tweets_total_keywords = sheet.max_row
    # column_count = sheet.max_column

# start charts from here
    dataSource = OrderedDict()
    # The `chartConfig` dict contains key-value pairs of data for chart attribute
    chartConfig = OrderedDict()
    chartConfig["caption"] = "TWEETS BY SENTIMENT"
    chartConfig["theme"] = "fusion"

    dataSource["chart"] = chartConfig
    dataSource["data"] = []
    dataSource["data"].append({"label": 'Total Tweets', "value": tweets_count})
    dataSource["data"].append(
        {"label": 'Total Users', "value": count_username})
    dataSource["data"].append(
        {"label": 'Total Location', "value": tweets_location})
    dataSource["data"].append(
        {"label": 'Key Words Define', "value": tweets_total_keywords})
    pie3d = FusionCharts("pie3d", "ex2", "100%", "400", "chart-1", "json",
                         # The data is passed as a string in the `dataSource` as parameter.

                         {
                             "chart": {
                                 "caption": "TWEETS BY TYPE",
                                 # "subCaption" : "For a net-worth of $1M",
                                 "showValues": "1",
                                 "showPercentInTooltip": "0",
                                 # "numberPrefix" : "$",
                                 "enableMultiSlicing": "1",
                                 "theme": "fusion"
                             },
                             "data": [{
                                 "label": "Total Tweets",
                                 "value": tweets_count

                             }, {
                                 "label": "Total Users",
                                 "value": count_username
                             }, {
                                 "label": "Total Location",
                                 "value": tweets_location
                             }, {
                                 "label": "Key Words Define",
                                 "value": tweets_total_keywords

                             }]
                         })

    column2D = FusionCharts("column2d", "myFirstChart", "600",
                            "400", "myFirstchart-container", "json", dataSource)

    return render(request, 'blog/search.html', {
        'posts': posts, 'output': pie3d.render(), 'output1': column2D.render(), 'tweets_count': tweets_count, 'count_username': count_username, 'tweets_location': tweets_location, 'tweets_total_keywords': tweets_total_keywords
    })
    #twitter details
# Twiteer scraping
def twitter_search(request):
     return render(request, 'blog/twitter_search.html')
#twitter list page to show the tweets
def twitter_list(request):
       
    posts = []
    mydb._open_connection()
    search = request.GET.get('search')
    btn1 = request.GET.get('btn1')
    if search==None:
        sqlite_select_query = "SELECT * FROM tbl_twitter a JOIN tbl_hashtags b on a.Id=b.twitter_id ORDER BY a.Id DESC LIMIT 50"
        cursor.execute(sqlite_select_query)
    elif btn1=='tweets':
        sqlite_select_query = "SELECT * FROM tbl_twitter a JOIN tbl_hashtags b on a.Id=b.twitter_id WHERE b.title= %s ORDER BY a.Id DESC LIMIT 10"
        cursor.execute(sqlite_select_query,(search,))
   
    for row in cursor:
      
        posts.append({
            'text': row[4],
            'username': row[5],
            'dated': row[2],
            'retweetcount': row[3],
            'location': row[1],
            'urll': row[6],
            'description': row[7],
            'following': row[8],
            'followers': row[9],
            'tweetsdate':row[11]})
    mydb.commit()
    return render(request, 'blog/twitter_list.html', {
        'posts': posts })
   

def twitter_details(request):
        realtime = request.GET.get('realtime')
        database = request.GET.get('database')
        if realtime==None and database==None :
            return render(request, 'blog/index.html')

        else:
            if realtime=='realtime':
                search_words = request.GET.get('search')
                correntdatee = datetime.today().strftime('%Y-%m-%d')
                date_since=correntdatee
                numTweets=100
                numRuns=1
                rtotalTweets=0
                tweets_location=0
                retweets=0
                #text syntaments
                greatcounter = 0
                goodcounter = 0
                noutralcounter = 0
                badcounter = 0
                terriblecounter = 0
                total_followers=0
                posts=[]
                coutdate=[]
                for i in range(0, numRuns):
                    start_run = time.time()
                    tweets = tweepy.Cursor(api.search, q=search_words, lang="en", since=date_since, tweet_mode='extended').items(numTweets)
                    
                    tweet_list = [tweet for tweet in tweets]
                    noTweets = 0
                    for tweet in tweet_list:
                        rtotalTweets+=1
                        acctdesc = tweet.user.description
                        location = tweet.user.location
                        if location!='':
                            tweets_location+=1
                        following = tweet.user.friends_count
                        followers = tweet.user.followers_count
                        total_followers=total_followers+followers
                        totaltweets = tweet.user.statuses_count
                        usercreatedts = tweet.user.created_at
                        tweetcreatedts = tweet.created_at
                        coutdate.append({
                        #  t=count_date_row[1]
                        #  date.strftime('%m/%d/%Y')
                        'couttweets': rtotalTweets,
                        'bydate':tweetcreatedts.strftime('%m/%d/%Y')
                        })
                        retweetcount = tweet.retweet_count
                        retweets=retweets+retweetcount
                        # retweets=retweetcount+retweetcount
                        hashtags = tweet.entities['hashtags']
                        # tweet_id= tweet.id
                        username = tweet.user.screen_name
                        url =  f"https://twitter.com/user/status/{tweet.id}"
                        posts.append({
                        'urll': url})
                        try:
                            text = tweet.retweeted_status.full_text
                        except AttributeError:  # Not a Retweet
                            text = tweet.full_text
                        sentiments = sentiment_analysis(text)
                        if sentiments =='Great':
                            greatcounter+=1
                        if sentiments =='Good':
                            goodcounter+=1
                        if sentiments =='Neutral':
                            noutralcounter+=1
                        if sentiments =='Bad':
                            badcounter+=1
                        if sentiments =='Terrible':
                            terriblecounter+=1
                    return render(request, 'blog/index.html',{'totalTweets':rtotalTweets,'tweets_location':tweets_location,'retweets':retweets,'great':greatcounter,'good':goodcounter,'nutral':noutralcounter,'bad':badcounter,'terr':terriblecounter,'total_followers':total_followers,'posts': posts,'coutdate': json.dumps(coutdate)})
                    
            elif database=='database':
                posts = []
                mydb._open_connection()
                price_lte = request.GET.get('search')
                
                sqlite_select_query = "SELECT * FROM tbl_twitter a JOIN tbl_hashtags b on a.Id=b.twitter_id WHERE b.title= %s ORDER BY a.Id DESC LIMIT 50"
                cursor.execute(sqlite_select_query,(price_lte,))
                greatcounter = 0
                goodcounter = 0
                noutralcounter = 0
                badcounter = 0
                terriblecounter = 0
                for row in cursor:
                    sentiments = sentiment_analysis(row[4])
                    if sentiments =='Great':
                        greatcounter+=1
                    if sentiments =='Good':
                        goodcounter+=1
                    if sentiments =='Neutral':
                        noutralcounter+=1
                    if sentiments =='Bad':
                        badcounter+=1
                    if sentiments =='Terrible':
                        terriblecounter+=1
                    posts.append({
                        'text': row[4],
                        'username': row[5],
                        'dated': row[2],
                        'retweetcount': row[3],
                        'location': row[1],
                        'urll': row[6],
                        'description': row[7],
                        'following': row[8],
                        'followers': row[9],
                        'sentiment': sentiments})
                    
                mydb.commit()
                # show data from 2nd table
                tweets_countq_query = """select COUNT(distinct Username) as username,COUNT(a.Id),COUNT(distinct location) as location,SUM(retweetcount),COUNT(totaltweets),sum(followers) as Followers from tbl_twitter a JOIN tbl_hashtags b on a.Id=b.twitter_id WHERE b.title= %s ORDER BY a.Id"""
                cursor.execute(tweets_countq_query,(price_lte,))

                count_row = cursor.fetchone()
                count_username = count_row[0]
                tweets_count = count_row[1]
                tweets_location = count_row[2]
                retweets_count = count_row[3]
                total_tweets = count_row[4]
                total_followers = count_row[5]
                # end query
                coutdate = []
                cout_bydate = """select count(a.id),date(tweetcreatedts) as date FROM twitter.tbl_twitter a JOIN twitter.tbl_hashtags b on a.Id=b.twitter_id WHERE b.title=%s  
                group by date(tweetcreatedts)"""
                cursor.execute(cout_bydate,(price_lte,))
                for count_date_row in cursor:
                    coutdate.append({
                        'couttweets': count_date_row[0],
                        'bydate':count_date_row[1].strftime('%m/%d/%Y')
                        })
                mydb.commit()
                wb = xl.load_workbook("dictionary.xlsx", enumerate)
                sheet = wb.worksheets[0]

                tweets_total_keywords = sheet.max_row
                # column_count = sheet.max_column
                return render(request, 'blog/index.html', {
                    'posts': posts,'coutdate': json.dumps(coutdate),'tweets_count': tweets_count, 'count_username': count_username, 'tweets_location': tweets_location, 'tweets_total_keywords': tweets_total_keywords,
                    'great':greatcounter,'good':goodcounter,'nutral':noutralcounter,'bad':badcounter,'terr':terriblecounter,'retweets':retweets_count,'totalTweets':total_tweets,'total_followers':total_followers
                })
   
def mapper(request):
        #  search = 
         posts = []
        #  tweets = request.GET.get('tweets')
         if request.GET.get('search')==None:
             return render(request,'blog/mapper.html')
         else:
             search = request.GET.get('search')
             #this is used for tweets
             if request.GET.get('tweets')=='tweets':

                screen_name = request.GET.get('search')
                user = api.get_user(screen_name)
                name=user.name
                userName=user.screen_name
                location=user.location
                description=user.description
                website=user.url
                followers_count=user.followers_count
                image=user.profile_image_url
                totaltweets=user.statuses_count
                statuses = api.user_timeline(screen_name)
                # print(statuses)
                for status in statuses:
                    text=status.text
                    created=status.created_at
                    posts.append({
                        'text':text,
                        "created":created })
                return render(request,'blog/mapper.html',{'posts': posts,'totaltweets':totaltweets,'screenName':screen_name,'followerCount':followers_count,'location':location,'name':name,'image':image})
                #there are end of tweets

                #this is used for followers
             elif request.GET.get('followers')=='followers':
                followersArr=[]
                screen_name = request.GET.get('search')
                count=0
                for user in tweepy.Cursor(api.followers, screen_name).items(50): 
                    count += 1
                    print(count)
                    
                    followers_scrname=user.screen_name
                    followers_name=user.name
            
                    followersurl =  f"https://twitter.com/{user.screen_name}"
                    f_location=user.location
                    created=str(user.created_at)
                    followers_count=str(user.followers_count)
                    friend=str(user.friends_count)
                    image=user.profile_image_url
                    followersArr.append({
                        'followers_name':followers_name,
                        "followers_scrname":followers_scrname,
                        'followersurl':followersurl,
                        'f_location':f_location,
                        'created':created,
                        'followers_count':followers_count,
                        'friend':friend,
                        'image':image
                         })
                return render(request,'blog/mapper.html',{'followersArr': followersArr})
                #there are end of followers

            #this is used for followings
             elif request.GET.get('following')=='following':
                followingArr=[]

                screen_name = request.GET.get('search')
                count=0
                for user in tweepy.Cursor(api.friends, screen_name).items(50): 
                    count += 1

                    followers_scrname=user.screen_name
                    followers_name=user.name
                    followersurl =  f"https://twitter.com/{user.screen_name}"
                    f_location=user.location
                    created=str(user.created_at)
                    followers_count=str(user.followers_count)
                    friend=str(user.friends_count)
                    image=user.profile_image_url
                    
                    followingArr.append({
                        'following_name':followers_name,
                        "followers_scrname":followers_scrname,
                        'followingsurl':followersurl,
                        'f_location':f_location,
                        'created':created,
                        'following_count':followers_count,
                        'friend':friend,
                        'image':image

                         })
                print(followingArr)
                return render(request,'blog/mapper.html',{'followingArr': followingArr})
                #there are end of followings

def insertkeywords(request):

    keywords = request.GET.get("words")
    df = pd.read_excel('dictionary.xlsx')
    lstvalue = df['pk'].tail(1).index.item()
    df = df.append({'pk': lstvalue+5, 'words': keywords}, ignore_index=True)
    if keywords:
        df.to_excel('dictionary.xlsx')
    df = pd.read_excel('dictionary.xlsx')
    posts = []

    for index, row in df.iterrows():
        posts.append({
            'words': row['words'],
            'pk': row['pk'],

        })

    # items=df
    context = {
        'items': posts,
        'header': 'Laptops',
    }

    return render(request, 'blog/dictionary.html', context)

   
